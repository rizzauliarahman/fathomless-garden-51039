import numpy as np
import openpyxl as pyx
import random
import os
import pickle
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient import errors
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import json


def load_dataset():
    x_train = []
    y_train = []

    x_class, y_class = load_class("Training_Raw_Data_Class_1.xlsx", 0)
    x_train.extend(x_class)
    y_train.extend(y_class)

    x_class, y_class = load_class("Training_Raw_Data_Class_2.xlsx", 1)
    x_train.extend(x_class)
    y_train.extend(y_class)

    dataset = []

    for x, y in zip(x_train, y_train):
        dataset.append([x, y])

    random.shuffle(dataset)

    x_train = [dataset[i][0] for i in range(len(dataset))]
    y_train = [dataset[i][1] for i in range(len(dataset))]

    label_list = ["Class 1", "Class 2"]

    data = dict()
    data['x_train'] = np.array(x_train)
    data['y_train'] = np.array(y_train, dtype=int)
    #
    # dataset = dict()
    # dataset["data"] = data
    # dataset["label_list"] = label_list
    #
    # svfile = open("dataset.dat", mode="wb")
    # pickle.dump(dataset, svfile)

    return data, label_list


def load_new_data():
    x_add, y_add = check_new()

    return x_add, y_add


def load_class(filename, class_label):
    wb = pyx.load_workbook(filename)
    wss = wb.worksheets

    x_class = []
    y_class = []

    for i in range(2, wss[1].max_row + 1):
        data = []
        for j in range(2, wss[1].max_column):
            pair = [wss[1].cell(row=i, column=j).value, wss[2].cell(row=i, column=j).value,
                    wss[3].cell(row=i, column=j).value]
            data.append(pair)

        x_class.append(data)
        y_class.append(class_label)

    return x_class, y_class


def check_new():
    SCOPES = ['https://www.googleapis.com/auth/drive']

    creds = None

    if os.path.exists("token.pickle"):
        with open("token.pickle", mode="rb") as token:
            creds = pickle.load(token)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
            creds = flow.run_local_server()

        with open("token.pickle", mode="wb") as token:
            pickle.dump(creds, token)

    service = build("drive", "v2", credentials=creds)

    folder_id = "1tN8U5L1pspnhZuCoJht7yW7FTRsE5-WJ"

    results = service.children().list(folderId=folder_id).execute()
    items = results.get("items", [])

    x_add = []
    y_add = []

    for item in items:
        file = service.files().get(fileId=str(item['id'])).execute()
        if file['mimeType'] == 'application/json':
            x_add.append(parse_json(service, file))
        elif file['title'] == 'input_result.txt':
            y_add = parse_labels(service, file)
        service.files().delete(fileId=str(item['id'])).execute()

    return x_add, y_add


def parse_json(service, file):
    request = service.files().get_media(fileId=file['id'])
    svfile = open(file['title'], mode="wb")
    media_request = MediaIoBaseDownload(svfile, request)

    download_progress, done = media_request.next_chunk()
    svfile.close()

    with open(file['title'], mode="r") as jsfile:
        data = json.load(jsfile)

    x = []
    for val in data['valueList']:
        raw = val['input']

        d = [raw['Chart_1'], raw['Chart_2'], raw['Chart_3']]
        x.append(d)

    jsfile.close()
    os.remove(file['title'])

    return x


def parse_labels(service, file):
    request = service.files().get_media(fileId=file['id'])
    svfile = open(file['title'], mode="wb")
    media_request = MediaIoBaseDownload(svfile, request)

    download_progress, done = media_request.next_chunk()
    svfile.close()

    y = []
    with open(file['title'], mode="r") as labelfile:
        line = labelfile.readline()

        while line:
            y.append(int(line))
            line = labelfile.readline()

    labelfile.close()
    os.remove(file['title'])

    return y
